from dataclasses import dataclass

from pickle import FALSE
from django.shortcuts import render
from .models import store
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .forms import UserForm, UserProfileInfoForm, UpdateProfileForm
from django.views.generic import DetailView,UpdateView,TemplateView
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .models import UserProfileInfo, User
from django.shortcuts import get_object_or_404
import random
import sklearn
from sklearn.metrics.pairwise import cosine_similarity
from textblob import TextBlob
from textblob.sentiments import *
from nltk.stem.wordnet import WordNetLemmatizer
lmtzr = WordNetLemmatizer()
from sklearn.feature_extraction.text import TfidfVectorizer
import nltk 
import string 
import re
from nltk.corpus import stopwords 
from nltk.tokenize import word_tokenize
nltk.download('punkt')
from bs4 import BeautifulSoup
nltk.download('stopwords')
import string #has the list of all punctuations
from nltk.stem import WordNetLemmatizer
nltk.download('wordnet')
import pandas as pd
lemmer = nltk.stem.WordNetLemmatizer()
stop_w = stopwords.words('english')



GREETING_INPUTS = ("hello", "hi", "greetings", "sup", "what's up","hey",)
GREETING_RESPONSES = ["hi", "hey", "*nods*", "hi there", "hello", "I am glad! You are talking to me"]
greetings = ['hi','hey', 'hello', 'heyy', 'hi', 'hey', 'good evening', 'good morning', 'good afternoon', 'good', 'fine', 'okay', 'great', 'could be better', 'not so great', 'very well thanks', 'fine and you', "i'm doing well", 'pleasure to meet you', 'hi whatsup']
happy_emotions = ['i feel good', 'life is good', 'life is great', "i've had a wonderful day", "i'm doing good"]
goodbyes = ['thank you', 'thank you', 'yes bye', 'bye', 'thanks and bye', 'ok thanks bye', 'goodbye', 'see ya later', 'alright thanks bye', "that's all bye", 'nice talking with you', 'i’ve gotta go', 'i’m off', 'good night', 'see ya', 'see ya later', 'catch ya later', 'adios', 'talk to you later', 'bye bye', 'all right then', 'thanks', 'thank you', 'thx', 'thx bye', 'thnks', 'thank u for ur help', 'many thanks', 'you saved my day', 'thanks a bunch', "i can't thank you enough", "you're great", 'thanks a ton', 'grateful for your help', 'i owe you one', 'thanks a million', 'really appreciate your help', 'no', 'no goodbye']

remove_punct_dict = dict((ord(punct), None) for punct in string.punctuation)
@login_required
def user_logout(request):
    #del request.session['user_id']
    request.session.flush()
    logout(request)
    # return HttpResponseRedirect(reverse('user_login'))
    # return HttpResponseRedirect('/user_login')
    return render(request, 'login.html')

def login(request):
    return render(request, 'login.html')

class Aboutpageview(TemplateView):
    template_name=' ';

def register(request):

    registered = False

    if request.method == 'POST':
        user_form = UserForm(data=request.POST)
        profile_form = UserProfileInfoForm(data=request.POST)

        if user_form.is_valid() and profile_form.is_valid():

            user = user_form.save()
            user.set_password(user.password)
            user.save()

            profile = profile_form.save(commit=False)
            profile.user = user

            if 'profile_pic' in request.FILES:
                profile.profile_pic = request.FILES['profile_pic']

            profile.save()

            registered = True

        else:
            print(user_form.errors, profile_form.errors)

        if registered:
            # return HttpResponseRedirect('/user_login')
            return render(request, 'bmi.html')

    else:
        user_form = UserForm()
        profile_form = UserProfileInfoForm()

    return render(request, 'register.html',
                  {'user_form': user_form, 'profile_form': profile_form, 'registered': registered})


def user_login(request):

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:
            if user.is_active:
                # login(request, user)
                # request.session['user_id'] = user.profile.pk

                # return HttpResponseRedirect(reverse('predict:predict', kwargs={'pk': user.profile.pk}))
                return render(request, 'bmi.html')
            else:
                return HttpResponse("Account not active")
        else:
            print("Tried login and failed")
            print("username: {} and password: {}".format(username, password))
            return HttpResponse("Invalid login details supplied!")

    else:
        return render(request, 'login.html', {})

class ProfileDetailView(LoginRequiredMixin, DetailView):
    login_url = '/'
    redirect_field_name = '/'
    model = UserProfileInfo
    template_name = 'profileview.html'

    def get_context_data(self, **kwargs):
        if self.request.session.has_key('user_id'):
            u_id = self.request.session['user_id']
            context = super(ProfileDetailView, self).get_context_data(**kwargs)
            context['user_id'] = u_id

        return context




def clean(column,df,stopwords=False):
  #remove stop words
  df[column] = df[column].apply(str)
  df[column] = df[column].str.lower().str.split()
  if stopwords:
        df[column]=df[column].apply(lambda x: [item for item in x if item not in stop_w])
  #remove punctuation
  df[column]=df[column].apply(lambda x: [item for item in x if item not in string.punctuation])
  df[column]=df[column].apply(lambda x: " ".join(x))

def LemTokens(tokens):
    return [lemmer.lemmatize(token) for token in tokens]

def LemNormalize(text):
    text = text.lower()
    clean_text = text.translate(remove_punct_dict)
    words = nltk.word_tokenize(clean_text)
    return LemTokens(words)
def parents():
    df = pd.read_csv("parents.csv",encoding='latin-1')
    df2 = df
    

    clean('questionText',df2)
    df2=df2.fillna(0)

    questions = df2["questionText"].to_list()

    answers = df2["answerText"].to_list()
    sent_tokens =questions
    ans_sent_tokens=answers

    
    # answers=re.sub("[^a-zA-Z]", " ",str(answers))
    # questions=re.sub("[^a-zA-Z]", " ",str(questions))
    ques=[]
    ans=[]
    for i in questions:
        ques.append(i)
    for j in answers:
        ans.append(j)

    sent_tokens =ques
    ans_sent_tokens=ans



    return ques,ans_sent_tokens,sent_tokens

def greeting(sentence):
    for word in sentence.split():
        if word.lower() in GREETING_INPUTS:
            return random.choice(GREETING_RESPONSES)
def response(user_response,sent_tokens,ans_sent_tokens):
    robo_response=''
    sent_tokens.append(user_response)
    TfidfVec = TfidfVectorizer(tokenizer=LemNormalize, stop_words='english')
    tfidf = TfidfVec.fit_transform(sent_tokens)
    vals = cosine_similarity(tfidf[-1], tfidf)
    idx=vals.argsort()[0][-2]
    # print(sent_tokens)
    print(len(sent_tokens))
    flat = vals.flatten()
    flat.sort()
    req_tfidf = flat[-2]
    print(req_tfidf)
    if(req_tfidf>=0.5):
        robo_response = robo_response+sent_tokens[idx]
        if idx<(len(sent_tokens)-1):
            print(ans_sent_tokens[idx])
            return ans_sent_tokens[idx]
        
    else:
        robo_response=robo_response+"I am sorry! I don't understand you"
        return robo_response

def data(userText):
    
    flag=True

    while(flag==True):
        # user_response = input()
        user_response=userText.lower()
        if(user_response!='bye'):
            if(user_response=='thanks' or user_response=='thank you' ):
                flag=False
                return "You are welcome.."
                # print("BOT: You are welcome..")
            else:

                if(greeting(user_response)!=None):
                    return greeting(user_response)
                else:
                    # print("BOT: ",end="")
                    # if who =="par":
                    ques,ans_sent_tokens,sent_tokens=parents()
                    # else:
                    #     ques,ans_sent_tokens,sent_tokens=stud()
                    res=response(user_response,ques,ans_sent_tokens)
                    sent_tokens.remove(user_response)
                    return res
        else:
            flag=False
            return " Bye! take care.."
            # return JsonResponse({'status':'OK','answer':"Bye! take care.."})
# Create your views here.
@login_required(login_url='/')
def bmi(request):
    return render(request,"bmi.html")
@csrf_exempt
def chat(request):
    return render(request,"index.html")

def bmi_cal(request):
    if request.method == 'POST':
        height = int(request.POST.get('height'))
        weight = int(request.POST.get('weight'))
    h=(height/100)
    h=h**2
    bmi = round((weight/h),2)
    context={"bmi":bmi}
    print(bmi)
    return render(request,"bmi.html",context)
    
    
def insert(request):
    # try:
    #     bot=request.POST.get('messageText')
    #     user = store.objects.filter(user = bot).values_list("response", flat=True)
    #     bot_response =user[0]
    #     bot_response = str(bot_response)      
    #     print(bot_response)
    #     return JsonResponse({'status':'OK','answer':bot_response})
    # except:
    #     #user1="sorry i don't understand "
    #     bot_response = "sorry i don't understand "      
    #     print(bot_response)
    #     return JsonResponse({'status':'OK','answer':bot_response})


    message1 = request.POST.get('messageText')
    print(message1)
    message=message1.lower()
    val=0
    if message.isnumeric():
        val=int(message)
        print(int(message))
    import json
    import requests
    
    try:
        api_url='https://api.api-ninjas.com/v1/nutrition?query='
        api_request=requests.get(api_url + message ,headers= {'X-Api-Key': '5N7iKNAXiJy7N2KjpnKWXg==AH2UJ6cd29gWhXLW'})
        api=json.loads(api_request.content)
        print(api_request.content)
        
    except Exception as e:
        # api="oops! There was an error"
        print(e)
        
    
    # bot_response=database(message)

    # bot_response =  bot.get_response(message)                                                                                
    # reply = "i need some product"
    # bot_response=reply
    # print(bot_response)
    global height
    
    global bmi
    while True:
        blob = TextBlob(message, analyzer=PatternAnalyzer())
        polarity = blob.sentiment.polarity

        if message in greetings:
            # return "Hello! How may I help you today?"
            return JsonResponse({'status':'OK','answer':"Hello! How may I help you today?"})
        elif polarity>0.7:
            # return "That's great! Do you still have any questions for me?"
            return JsonResponse({'status':'OK','answer':"That's great! Do you still have any questions for me?"})
        elif message in happy_emotions:
            # return "That's great! Do you still have any questions for me?" 
            return JsonResponse({'status':'OK','answer':"That's great! Do you still have any questions for me?"}) 
        elif message in goodbyes:
            # return "Hope I was able to help you today! Take care, bye!"
            return JsonResponse({'status':'OK','answer':"Hope I was able to help you today! Take care, bye!"}) 
        


        # if message == ("bye"):

        #     bot_response='Hope to see you soon'

        #     print(bot_response)
        #     return JsonResponse({'status':'OK','answer':bot_response})

        elif "bmi" in message:

            bot_response='what is your height in cm?'     
            print(bot_response)
            bmi=0
            return JsonResponse({'status':'OK','answer':bot_response})
        elif val > 120 and val < 200 :
            
            bot_response='what is your weight in Kg?'     
            print(bot_response)
            height=val
            bmi=1
            return JsonResponse({'status':'OK','answer':bot_response})
        
        elif val > 20 and val < 200 :
            if bmi:
                h=(height/100)
                h=h**2
                bmi = round((val/h),2)
                bot_response="Your bmi is "+str(bmi)
                weight='your is under weight'
                print(height)     
                print(bot_response)
                weight=val
                # bmi=0
                if bmi < 18.5:
                    content="you are underweight, it is important to eat a variety of foods that give you the nutrition you need. You should make sure you eat enough energy to gain weight, protein to repair your body and build your muscles, and vitamins and minerals to make you healthy"
                elif bmi>18.5 and bmi<23.0:
                    content="your totally normal"
                elif bmi >25:
                    content="you are overweight,Obesity is a complex disease involving an excessive amount of body fat. Obesity isn't just a cosmetic concern. It's a medical problem that increases the risk of other diseases and health problems, such as heart disease, diabetes, high blood pressure and certain cancers."



                return JsonResponse({'status':'OK','answer':bot_response,'bmi':content})
        
        elif api:
            # print(type(api[0]))
            api=api[0]
            # name=list(api.keys())
            b="\t the neutritian value of \n " +str(api.get('name'))+" is"+str(api)
            
            return JsonResponse({'status':'OK','answer':str(b)})


            

        else:
        
            try:
                
                user = store.objects.filter(user = message).values_list("response", flat=True)
                bot_response =user[0]
                bot_response = str(bot_response)      
                print(bot_response)
                return JsonResponse({'status':'OK','answer':bot_response})
            except:
                #user1="sorry i don't understand "
                topic = data(message1)
                print (topic)
                # res = random.choice(dictionary[topic])
                # print (res)
                if topic:
                    return JsonResponse({'status':'OK','answer':topic}) 
                else:
                    return JsonResponse({'status':'OK','answer':"I am sorry! I don't understand you"}) 
                # return JsonResponse({'status':'OK','answer':bot_response})
        
        # else: 
            
            # return topic
def input(request):
    return render(request,'commands.html')
def addcommant(request):
    catch=store()
    catch.user=request.POST.get('user')
    catch.response=request.POST.get('bot')
    catch.save()
    return render(request,'index.html')

def food(request):
    return render(request,'food.html')

def diet(request):
    weight=request.POST.get('diet')
    if weight == 'overweight':   
        return render(request,'over.html')
    else:
        return render(request,'under.html')
    
def img(request):
    from django.core.files.storage import FileSystemStorage
    from tensorflow.keras.preprocessing import image
    from openpyxl.reader.excel import load_workbook
    import tensorflow as tf
    from django.core.files.storage import default_storage
    from tensorflow import keras
    import numpy as np
    import cv2
    from PIL import Image as img
    import random
    import os
    
    if request.method == 'POST':
        batch_size = 32
        img_height = 64
        img_width = 64
        dir = './static/img/upload'
        for f in os.listdir(dir):
            os.remove(os.path.join(dir, f))
        model_dl = keras.models.load_model("model.h5") 
        
        dict = {0:'Apple pie',1:'Baby back ribs',2:'Baklava',3:'Beef carpaccio',4:'Beef tartare',5:'Beet salad',6:'Beignets',7:'Bibimbap',8:'Bread pudding',9:'Breakfast burrito',
        10:'Bruschetta',11:'Caesar salad',12:'Cannoli',13:'Caprese salad',14:'Carrot cake',15:'Ceviche',16:'Cheesecake',17:'Cheese plate',18:'Chicken curry',19:'Chicken quesadilla',
        20:'Chicken wings',21:'Chocolate cake',22:'Chocolate mousse',23:'Churros',24:'Clam chowder',25:'Club sandwich',26:'Crab cakes',27:'Creme brulee',28:'Croque madame',
        29:'Cup cakes',30:'Deviled eggs',31:'Donuts',32:'Dumplings',33:'Edamame',34:'Eggs benedict',35:'Escargots',36:'Falafel',37:'Filet mignon',38:'Fish and chips',
        39:'Foie gras',40:'French fries',41:'French onion soup',42:'French toast',43:'Fried calamari',44:'Fried rice',45:'Frozen yogurt',46:'Garlic bread',47:'Gnocchi',
        48:'Greek salad',49:'Grilled cheese sandwich',50:'Grilled salmon',51:'Guacamole',52:'Gyoza',53:'Hamburger',54:'Hot and sour soup',55:'Hot dog',56:'Huevos rancheros',
        57:'Hummus',58:'Ice cream',59:'Lasagna',60:'Lobster bisque',61:'Lobster roll sandwich',62:'Macaroni and cheese',63:'Macarons',64:'Miso soup',65:'Mussels',
        66:'Nachos',67:'Omelette',68:'Onion rings',69:'Oysters',70:'Pad thai',71:'Paella',72:'Pancakes',73:'Panna cotta',74:'Peking duck',75:'Pho',76:'Pizza',77:'Pork chop',
        78:'Poutine',79:'Prime rib',80:'Pulled pork sandwich',81:'Ramen',82:'Ravioli',83:'Red velvet cake',84:'Risotto',85:'Samosa',86:'Sashimi',87:'Scallops',88:'Seaweed salad',
        89:'Shrimp and grits',90:'Spaghetti bolognese',91:'Spaghetti carbonara',92:'Spring rolls',93:'Steak',94:'Strawberry shortcake',95:'Sushi',96:'Tacos',97:'Takoyaki',
        98:'Tiramisu',99:'Tuna tartare',100:'Waffles'}
        img = request.FILES['img1']
        # file=img.save(filename='1.jpg')
        # print(img)
        # image = st.file_uploader(img, type=["JPEG", "JPG", "PNG"])
        file_name = "static/img/upload/pic.png"
        default_storage.save(file_name, img)
        img_to_detect =cv2.imread('static/img/upload/pic.png', cv2.IMREAD_COLOR)
        # img=cv2.resize(img, (img_rows, img_cols))
        # img = cv2.cvtColor(cv2.imread('lena_caption.png'), cv2.COLOR_BGR2RGB)
        # if(type(image) == type(None)):
        #     pass
        # else:
        # print(len(image))
        # cv2. imwrite('static/img/save.jpg',img_to_detect)
        
        img = cv2.resize(img_to_detect, (64,64), interpolation=cv2.INTER_AREA)
        # img = cv2.resize(img_to_detect,(64,64))
        x = image.img_to_array(img) 
        x = np.expand_dims(x, axis=0)
        imag = np.vstack([x])
        classes = model_dl.predict_classes(imag, batch_size=batch_size)
        text = str(dict[classes.item()])
        print(text)
        wrkbk = load_workbook(r"calorie.xlsx")
        sh = wrkbk.active
        
        for i in range(1,101):
            c=str(sh.cell(row=i,column=1).value)
            # print(c)
            if c == text:
                calo =str(sh.cell(row=i,column=2).value)
                print("calories",calo)
                typr =str(sh.cell(row=i,column=4).value)
                if typr == 'yes':
                    print("Its a junk food-")
                elif typr == 'no':    
                    print("Its a healthy food")
        # cv2.putText(img_to_detect,text,(45,60),cv2.FONT_HERSHEY_SIMPLEX,1.25,(255,0,0),5) 
        # cv2.imshow("Detection Output", img_to_detect)
        import json
        import requests
        api_url='https://api.api-ninjas.com/v1/nutrition?query='
        api_request=requests.get(api_url + text ,headers= {'X-Api-Key': '5N7iKNAXiJy7N2KjpnKWXg==AH2UJ6cd29gWhXLW'})
        try:
            api=json.loads(api_request.content)
            # print(api_request.content)
        except Exception as e:
            api="oops! There was an error"
            print(e)
        # return render(request, 'home.html',{'api':api})
        context={"text":text,"calo":calo,"typr":typr ,"api":api}
        return render(request,'detail.html',context)
        